import tkinter as tk
from tkinter import simpledialog, messagebox
import asyncio
import threading
import qtm
import requests
import ntplib
from datetime import datetime, timedelta, timezone
import json
import os
import random
import openpyxl
from openpyxl import Workbook, load_workbook

# === CONFIGURATION ===
QTM_IP = "127.0.0.1"
AURORA_IP = "130.39.121.10"
VLC_TRIGGER_ENDPOINT = f"http://{AURORA_IP}:5000/start"
VIDEO_FOLDER_PATH = r"C:\\Users\\B24-Lab\\Desktop\\Videos"
NTP_SERVER = "time.google.com"

# === Globals ===
qtm_connection = None
loop = asyncio.new_event_loop()
session_data = []
played_videos = set()
trial_count = 0
current_trial = 0
patient_initials = ""
excel_file_path = ""
last_played_video = ""

# === Async Loop Setup ===
def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# === NTP SYNC ===
def sync_time_ntp(ntp_server):
    try:
        client = ntplib.NTPClient()
        response = client.request(ntp_server, version=3)
        ntp_time = datetime.fromtimestamp(response.tx_time)
        print(f"[✓] NTP Time Synced: {ntp_time}")
    except Exception as e:
        print(f"[!] NTP sync failed: {e}")

# === QTM Control ===
async def start_qtm_recording():
    global qtm_connection
    try:
        qtm_connection = await qtm.connect(QTM_IP)
        if qtm_connection is None:
            print("[!] Failed to connect to QTM.")
            return
        await qtm_connection.take_control("")
        await qtm_connection.start()
        print("[✓] QTM recording started.")
    except Exception as e:
        print(f"[!] QTM start error: {e}")

async def stop_qtm_recording():
    global qtm_connection
    try:
        if qtm_connection:
            await qtm_connection.stop()
            print("[✓] QTM recording stopped.")
            qtm_connection.disconnect()
            print("[✓] QTM disconnected.")
            qtm_connection = None
    except Exception as e:
        print(f"[!] QTM stop error: {e}")

# === Video Handling ===
def get_video_files():
    return [f for f in os.listdir(VIDEO_FOLDER_PATH)
            if f.endswith('.mp4') and os.path.isfile(os.path.join(VIDEO_FOLDER_PATH, f))]

def choose_video(category=None):
    global played_videos
    videos = get_video_files()
    if category:
        videos = [v for v in videos if category in v and v not in played_videos]
    else:
        videos = [v for v in videos if v not in played_videos]
    if not videos:
        return None
    return random.choice(videos)

def trigger_video_play(video_file, start_time):
    global last_played_video
    last_played_video = video_file
    try:
        payload = {"start_time": start_time.timestamp(), "video_file": video_file}
        response = requests.post(VLC_TRIGGER_ENDPOINT, json=payload)
        return response.status_code == 200
    except Exception as e:
        print(f"[!] Failed to trigger VLC: {e}")
        return False

def stop_video():
    try:
        requests.post(f"http://{AURORA_IP}:5000/stop")
    except Exception as e:
        print(f"[!] Failed to stop VLC: {e}")

# === Logging and Excel ===
def log_event(msg, start_time=None):
    now = datetime.now(timezone.utc)
    timestamp = now.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    delay = (now - start_time).total_seconds() if start_time else None
    print(f"[{timestamp}] {msg} | Delay: {delay:.3f} seconds" if delay else f"[{timestamp}] {msg}")
    session_data.append((msg, timestamp, f"{delay:.3f}" if delay else ""))
    return now

def save_to_excel():
    global excel_file_path, session_data, current_trial
    if not excel_file_path:
        today = datetime.now().strftime("%m_%d_%y")
        excel_file_path = f"{patient_initials}_{today}.xlsx"
        wb = Workbook()
    else:
        wb = load_workbook(excel_file_path)

    ws = wb.active
    row = ws.max_row + 2
    ws[f"A{row}"] = f"TRIAL {current_trial}"
    for i, (msg, ts, delay) in enumerate(session_data):
        ws[f"B{row+i}"] = msg
        ws[f"C{row+i}"] = ts
        ws[f"D{row+i}"] = delay

    wb.save(excel_file_path)
    session_data.clear()

# === GUI Functions ===
def on_start():
    global current_trial, trial_count, start_button
    if current_trial >= trial_count:
        messagebox.showinfo("Done", "All trials completed.")
        return
    current_trial += 1
    start_button.config(state='disabled')

    sync_time_ntp(NTP_SERVER)
    buffer_seconds = 5
    start_time = datetime.now(timezone.utc) + timedelta(seconds=buffer_seconds)
    log_event(f"Scheduled start time: {start_time}")

    asyncio.run_coroutine_threadsafe(start_qtm_recording(), loop)
    video = choose_video()
    if not video:
        messagebox.showerror("Error", "No available videos.")
        return

    played_videos.add(video)
    trigger_video_play(video, start_time)
    update_video_log(video)

def on_stop():
    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)
    stop_video()

    if messagebox.askyesno("Save Data", "Save timestamps and delay data to Excel?"):
        save_to_excel()

def on_replay():
    if last_played_video:
        start_time = datetime.now(timezone.utc) + timedelta(seconds=3)
        trigger_video_play(last_played_video, start_time)
        log_event(f"Replaying video: {last_played_video}")

def on_next(category):
    global current_trial
    if current_trial >= trial_count:
        messagebox.showinfo("Done", "All trials completed.")
        return
    current_trial += 1
    video = choose_video(category)
    if not video:
        messagebox.showerror("Error", f"No more {category} videos left.")
        return
    played_videos.add(video)
    start_time = datetime.now(timezone.utc) + timedelta(seconds=3)
    trigger_video_play(video, start_time)
    update_video_log(video)

def update_video_log(video):
    if "Social" in video:
        tag = "Social"
    elif "NonSocial" in video:
        tag = "NonSocial"
    else:
        tag = "Unknown"
    video_log.insert(tk.END, f"Trial {current_trial}: {video} ({tag})")
    video_log.see(tk.END)

# === Initialization ===
def initialize():
    global trial_count, patient_initials
    root = tk.Tk()
    root.withdraw()
    trial_count = simpledialog.askinteger("Trial Count", "Enter number of trials:", minvalue=1)
    patient_initials = simpledialog.askstring("Patient Initials", "Enter patient initials:")
    root.destroy()

# === GUI ===
def launch_gui():
    global start_button, video_log

    root = tk.Tk()
    root.title("Sync Controller")
    root.geometry("500x400")

    tk.Label(root, text="Trial Controller", font=("Helvetica", 16)).pack(pady=10)

    start_button = tk.Button(root, text="START", font=("Helvetica", 12), bg="green", fg="white", command=on_start)
    start_button.pack(pady=5)

    tk.Button(root, text="STOP", font=("Helvetica", 12), bg="red", fg="white", command=on_stop).pack(pady=5)
    tk.Button(root, text="REPLAY VIDEO", font=("Helvetica", 12), command=on_replay).pack(pady=5)
    tk.Button(root, text="NEXT - SOCIAL", font=("Helvetica", 12), command=lambda: on_next("Social")).pack(pady=5)
    tk.Button(root, text="NEXT - NONSOCIAL", font=("Helvetica", 12), command=lambda: on_next("NonSocial")).pack(pady=5)

    video_log = tk.Listbox(root, width=70)
    video_log.pack(pady=10)

    root.mainloop()

# === MAIN ===
if __name__ == "__main__":
    initialize()
    launch_gui()
